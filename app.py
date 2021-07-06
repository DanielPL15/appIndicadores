# -*- coding: utf-8 -*-
"""
Created on Tue Mar 23 17:37:44 2021

@author: Daniel Peña
"""


from functionalities.draw_vertex3d import draw_vertex3d
from functionalities.draw_plot2d import plot_2d
from others.read_indicator_sunburst import read_indicator_sunburst
from others.read_indicator import read_indicator
from dash_html_components.Div import Div
from dash_html_components.Spacer import Spacer
import pandas as pd
import dash_bootstrap_components as dbc
import dash
import dash_core_components as dcc
import dash_html_components as html
from dash.dependencies import ALL, State, MATCH, ALLSMALLER, Input, Output, State
import plotly.express as px
from functionalities.draw_radar import draw_radar
from functionalities.draw_parallel import draw_parallel
from functionalities.draw_sunburst import draw_sunburst
from functionalities.draw_map import draw_map
from functionalities.draw_vertex import draw_vertex
# from functionalities.draw_vertex import draw_vertex
import dash_auth
import xlrd
import openpyxl
from others.write_excel import *
from others.write_excel2 import *
from others.deleteSheet import *
import glob
import ntpath
import random
import dash_table
from whitenoise import WhiteNoise
import numpy as np
from sklearn.cluster import KMeans
from others.deleteColumns import *

#df = pd.read_excel("indicators/ODSy3i.xlsx", engine='openpyxl')

file2='static/filtros.xlsx'
df1 = pd.read_excel(file2, sheet_name='Hoja1', engine='openpyxl')
df2 = pd.read_excel(file2, sheet_name='Cluster', engine='openpyxl')
df3 = pd.read_excel(file2, sheet_name='Parallel', engine='openpyxl')
df_list_groups_of_countries = pd.read_excel('static/Grupos_paises.xlsx',sheet_name='Hoja1', engine='openpyxl')

# Creation of a list of dictionaries. Each of them with the name and ID of one country
df6 = df_list_groups_of_countries[['ISO_NUESTRO','country']]
df6.rename(columns={'ISO_NUESTRO' : 'value', 'country' : 'label'}, inplace = True)
df6_dict = df6.to_dict('records')


df_types_regressions=[{'label':'linear','value':'linear'},
                      {'label':'exponential','value':'exponential'},
                      {'label':'poly2','value':'poly2'},
                      {'label':'poly3','value':'poly3'},
                      {'label':'logarithmic','value':'logarithmic'},
                      {'label':'logistic','value':'logistic'},
                      {'label':'all','value':'all'}]

# Names of the excel files
file_list_path = glob.glob("indicators/*.xlsx")
n=0
file_list=[]
for i in file_list_path:
    file_list.append(ntpath.basename(i))
    n=n+1



# Default values for dropdows
df_file_list = pd.DataFrame({'label':file_list,'value':file_list_path})
df_file_list_dict = df_file_list.to_dict('records')

# Default values for dropdows of group of countries
data_top = df_list_groups_of_countries.head(1)
var=data_top.iloc[1:1, 3:32].columns
df_list_groups_of_countries_2col = pd.DataFrame({'label':var,'value':var})
df_list_groups_of_countries_dict = df_list_groups_of_countries_2col.to_dict('records')


book = openpyxl.load_workbook("indicators\ODSy3i.xlsx")
default_sheets = book.sheetnames
df_default_sheets = pd.DataFrame({'label':default_sheets,'value':default_sheets})
df_default_sheets_dict = df_default_sheets.to_dict('records')
df_aux = pd.read_excel("indicators\ODSy3i.xlsx", sheet_name='ODS', engine='openpyxl')
default_indicators = list(df_aux.columns)[2:]
df_default_indicators = pd.DataFrame({'label':default_indicators,'value':default_indicators})
df_default_indicators_dict = df_default_indicators.to_dict('records')

df_table_clusters=pd.DataFrame()


# DataFrame with file, sheet and indicator of cluster variables
df_cluster_vars = pd.DataFrame()

# Dictionary with the variables for the cluster: contains file, sheet and indicator name
cluster_variables_dict={}
cluster_variables_list=[]


options_method_vertex = pd.DataFrame({'label':['Auto','Ball Tree','KD tree','Brute'],'value':['auto','ball_tree','kd_tree','brute']})
options_method_vertex = options_method_vertex.to_dict('records')

options_neigh_vertex = pd.DataFrame({'label':['1','2','3','4','5','6','7'],'value':['1','2','3','4','5','6','7']})
options_neigh_vertex = options_neigh_vertex.to_dict('records')

count_number_of_list_cluster=pd.DataFrame(['0'],columns=list('n')) # counts the number of clusters


# you need to include __name__ in your Dash constructor if
# you plan to use a custom CSS or JavaScript in your Dash apps
app = dash.Dash(__name__,  external_stylesheets=[dbc.themes.SUPERHERO], suppress_callback_exceptions=True)


server = app.server

server.wsgi_app = WhiteNoise(server.wsgi_app, root='static/')

auth = dash_auth.BasicAuth(
    app,
    {'dani': '1234',
    'josefelix':'1234'}
    )   


# Content of Filter Tab
def filter_tab():

    return html.Div([
        dbc.Row(
            [
                dbc.Col(
                    [
                        html.Div([
                            html.H5('Considered countries', style={'align':'left'}),
                            html.Button('Save', id='save_considered', className='btn btn-secondary', n_clicks=0),
                            dbc.Toast(
                                "Saved correctly",
                                id="toast_saved",
                                is_open=False,
                                dismissable=True,
                                duration=2500,
                                icon="success"),
                        ], style = {'padding':'20px','text-align':'left'})
                    ],
                    width={'size': 3, 'offset': 0}
                ),
                dbc.Col(
                    [   
                        html.Div([
                            html.H5('Desired country labels', style={'align':'left'}),
                            html.Button('Save', id='save_desired', className='btn btn-secondary', n_clicks=0),
                            dbc.Toast(
                                "Saved correctly",
                                id="toast_desired",
                                is_open=False,
                                dismissable=True,
                                duration=2500,
                                icon="success"),
                        ],style = {'text-align':'left'}),
                    ],
                    width={'size': 3, 'offset': 3}
                )
            ],align='start'),
        dbc.Row(
            [
                dbc.Col(
                        dbc.Checklist(id='ckl_considered',
                                    options= df6_dict,
                                    className='container p-3 my-3 border'
                                    ),
                        width={'size': 3, 'offset': 0},
                ),
                dbc.Col([
                        html.H6('Number of selected countries: 0', id='number_countries_considered'),
                        html.H5('Filter by group'),
                        dcc.Dropdown(
                            id='dropdown_groups_countries',
                            options=df_list_groups_of_countries_dict,
                            className="dropbtn",
                            value= df_list_groups_of_countries_2col.at[1,'value'],
                            placeholder='Select a group of indicators',
                            clearable=False,
                        ),
                        html.Div(
                                [
                                html.Div([
                                    html.Button('Include All', id='include_all_considered', className='btn btn-secondary', n_clicks=0),
                                ],style={'padding':'5px'}),
                                html.Div([
                                    html.Button('Clear All', id='clear_all_considered', className='btn btn-secondary', n_clicks=0),
                                ],style={'padding':'5px'}),
                                ], style = {'padding':'5px'}),
                ],
                    width={'size': 3, 'offset': 0},
                ),
                dbc.Col(
                        dbc.Checklist(id='ckl_desired',
                                    options= df6_dict,
                                    className='container p-3 my-3 border'
                                    ),
                        width={'size': 3, 'offset': 0},
                ),
                dbc.Col([
                        html.H6('Number of selected countries: 0', id='number_countries_desired'),
                        html.H5('Filter by group'),
                        dcc.Dropdown(
                            id='dropdown_groups_countries_labels',
                            options=df_list_groups_of_countries_dict,
                            className="dropbtn",
                            value= df_list_groups_of_countries_2col.at[1,'value'],
                            placeholder='Select a group of indicators',
                            clearable=False,
                        ),
                        html.Div(
                                [
                                html.Div([
                                    html.Button('Include All', id='include_all_desired', className='btn btn-secondary', n_clicks=0),
                                ],style={'padding':'5px'}),
                                html.Div([
                                    html.Button('Clear All', id='clear_all_desired', className='btn btn-secondary', n_clicks=0),
                                ],style={'padding':'5px'}),
                                ], style = {'padding':'5px'}),
                    ],
                    width={'size': 3, 'offset': 0},
                )                
            ],
         style={'padding':'20px','align':'center'}),

    ])
# Content of Cluster Tab
def cluster_tab():
    deleteSheet(file2,'Cluster_var')
    count_number_of_list_cluster.iat[0,0]=0
    clear_dataframe = pd.DataFrame(columns=['var1','var2','var3','var4','var5','var6','var7','var8','var9','var10','var11','var12'])
    #clear_dataframe.loc[0]=0
    write_excel2(clear_dataframe,file2,"Cluster_var",0,0)
    return html.Div([
        dbc.Row([
            dbc.Col([
                html.H5('Select Cluster variables', style={'padding':'10px','background-color':'#06619e'}),
                dcc.Dropdown(
                    id='dropdown_file_cluster',
                    options=df_file_list_dict,
                    className="dropbtn",
                    value= "indicators\ODSy3i.xlsx",
                    placeholder='Select a group of indicators',
                    clearable=False,
                ),
                dcc.Dropdown(
                    id='dropdown_sheet_cluster',
                    options=df_default_sheets_dict,
                    className="dropbtn",
                    value='ODS',
                    placeholder='Select a subgroup of indicators',
                    clearable=False,
                ),
                dcc.Dropdown(
                    id='dropdown_indicator_cluster',
                    options=df_default_indicators_dict,
                    className="dropbtn",
                    value='2020 Global Index Score (0-100)',
                    placeholder='Select an indicator',
                    clearable=False,
                ),
                html.Div([
                    html.Div([
                        dbc.Button(
                            'Add Variable',
                            id="confirm_variable_cluster",
                            className='primary',
                            n_clicks=0,
                        )],
                         style={'padding':5,'width': '33%', 'display': 'inline-block'},
                        ),
                    html.Div([
                        dbc.Button(
                            'Clear List',
                            id="clear_variable_list",
                            className='primary',
                            n_clicks=0,
                        )],
                        style={'padding':5,'width': '33%', 'display': 'inline-block'},
                        ),
                    html.Div([
                        dbc.Button(
                            'Make Cluster',
                            id="make_cluster_button",
                            className='primary',
                            n_clicks=0,
                        )],
                        style={'padding':5,'width': '33%', 'display': 'inline-block'}
                    )],
                style={'padding':20}),
                html.Div([
                    html.H5('Number of Clusters', style={'padding':'10px','background-color':'#06619e'}),
                    dcc.Dropdown(
                                id='dropdown_number_clusters',
                                options=options_neigh_vertex,
                                className="dropbtn",
                                value = '3',
                                placeholder='Select Number of Clusters',
                                clearable=False,
                            ),
                    ]),
                ],
                width={'size': 4, 'offset': 0}),
            dbc.Col([
                html.H5('List of Cluster variables', style={'padding':'10px','background-color':'#06619e'}),
                dbc.ListGroup(
                    id='list_variables_cluster'
                    )
                ],   
                width={'size': 2, 'offset': 0}),
            dbc.Col([
                html.Div(
                        id = 'future_table_of_clusters',
                        children = []
                    )
                ],
                width={'size': 5, 'offset': 0}),
        ]),
    ], style={'padding':20})
# Content of Plot Tab
def plot_tab():
    return html.Div([
            html.H2('Plot a Graphic and scroll down to see it!', style={'padding':'10px','background-color':'#06619e', 'text-align':'center'}),
            dbc.Row([
                    dbc.Col([
                    html.Div(
                        id = 'menu_plot2D',
                            children = [html.H3('Plot a 2D graphic', style={'padding':'10px','background-color':'#06619e'}),
                            html.Div([
                                    html.H5('Variable in x', style={'padding':'10px','background-color':'#06619e'}),
                                    dcc.Dropdown(
                                        id='dropdown_file_x',
                                        options=df_file_list_dict,
                                        className="dropbtn",
                                        value= "indicators\ODSy3i.xlsx",
                                        placeholder='Select a group of indicators',
                                        clearable=False,
                                    ),
                                    dcc.Dropdown(
                                        id='dropdown_sheet_x',
                                        options=df_default_sheets_dict,
                                        className="dropbtn",
                                        value='ODS',
                                        placeholder='Select a subgroup of indicators',
                                        clearable=False,
                                    ),
                                    dcc.Dropdown(
                                        id='dropdown_indicator_x',
                                        options=df_default_indicators_dict,
                                        className="dropbtn",
                                        value='2020 Global Index Score (0-100)',
                                        placeholder='Select an indicator',
                                        clearable=False,
                                    ),
                            ],
                            style={'padding':5}),
                                html.Div([
                                    html.H5('Variable in y', style={'padding':'10px','background-color':'#06619e'}),
                                    dcc.Dropdown(
                                        id='dropdown_file_y',
                                        options=df_file_list_dict,
                                        className="dropbtn",
                                        value= "indicators\ODSy3i.xlsx",
                                        placeholder='Select a group of indicators',
                                        clearable=False,
                                    ),
                                    dcc.Dropdown(
                                        id='dropdown_sheet_y',
                                        options=df_default_sheets_dict,
                                        className="dropbtn",
                                        value='ODS',
                                        placeholder='Select a subgroup of indicators',
                                        clearable=False,
                                    ),
                                    dcc.Dropdown(
                                        id='dropdown_indicator_y',
                                        options=df_default_indicators_dict,
                                        className="dropbtn",
                                        value='2020 Global Index Score (0-100)',
                                        placeholder='Select an indicator',
                                        clearable=False,
                                    ),
                                    html.Div([
                                        dbc.Checklist(id='ckl_types_correlations',
                                                    options= df_types_regressions,
                                                    value = "linear",
                                                    className='container p-3 my-3 border',
                                                    labelStyle={'display': 'block'}
                                        ),
                                    ]),
                                    html.Div([
                                        dbc.Checklist(id='white_background',
                                                    options = [{'value':'White Background','label':'White Background'}])
                                    ]),
                                ],
                            style={'padding':5}),
                            html.Div(
                                html.Button('Plot', id='submit_plot2D', className='btn btn-secondary btn-lg btn-block', n_clicks=0),
                            style={'padding':5, 'textAlign': 'center'}),
                        ],
                        style={'padding':5}),
                    ],
                    width={'size': 3,  "offset": 0}),
                dbc.Col([
                    html.Div([
                        html.H3('Variables for Vertex', style={'padding':'10px','background-color':'#06619e'}),
                        dcc.Dropdown(
                            id='method_vertex',
                            options=options_method_vertex,
                            className="dropbtn",
                            value='auto',
                            placeholder='Select a method',
                            clearable=False,
                        ),
                        dcc.Dropdown(
                            id='dropdown_neighbors_vertex',
                            options=options_neigh_vertex,
                            className="dropbtn",
                            value='1',
                            placeholder='Select Number of Neighbors',
                            clearable=False,
                        ),
                        html.Div(
                        html.Button('Plot 2D', id='submit_vertex', className='btn btn-secondary btn-lg btn-block', n_clicks=0),
                        style={'padding':5, 'textAlign': 'center'}),
                        html.Div(
                        html.Button('Plot 3D', id='submit_vertex3d', className='btn btn-secondary btn-lg btn-block', n_clicks=0),
                        style={'padding':5, 'textAlign': 'center'}),
                    ],
                    style={'padding':5})
                ],
                width={'size': 3,  "offset": 0}),
                dbc.Col(
                        html.Div([
                            html.H3('Sunburst', style={'padding':'10px','background-color':'#06619e'}),
                            dcc.Dropdown(
                                id='dropdown_file_sunburst',
                                options=df_file_list_dict,
                                className="dropbtn",
                                value= "indicators\ODSy3i.xlsx",
                                placeholder='Select a group of indicators',
                                clearable=False,
                            ),
                            dcc.Dropdown(
                                id='dropdown_sheet_sunburst',
                                options=df_default_sheets_dict,
                                className="dropbtn",
                                value='ODS',
                                placeholder='Select a subgroup of indicators',
                                clearable=False,
                            ),
                            dcc.Dropdown(
                                id='dropdown_indicator_sunburst',
                                options=df_default_indicators_dict,
                                className="dropbtn",
                                value='2020 Global Index Score (0-100)',
                                placeholder='Select an indicator',
                                clearable=False,
                            ),
                            html.Div(
                            html.Button('Plot', id='submit_sunburst', className='btn btn-secondary btn-lg btn-block', n_clicks=0),
                            style={'padding':5, 'textAlign': 'center'}),
                        ],
                        style={'padding':5}),
                        width={'size': 3,  "offset": 0}),
                html.Div([
                    html.H3('Other graphics', style={'padding':'10px','background-color':'#06619e'}),
                    html.Div(
                    dbc.Col(html.Button('Map', id='submit_map', className='btn btn-secondary', n_clicks=0),
                        width={'size': 1,  "offset": 0}),
                    style={'padding':5}),
                    html.Div(
                    dbc.Col(html.Button('Parallel', id='submit_parallel', className='btn btn-secondary', n_clicks=0),
                        width={'size': 1,  "offset": 0}),
                    style={'padding':5}),
                    html.Div([
                    dbc.Col(html.Button('Radar', id='submit_radar', className='btn btn-secondary', n_clicks=0),
                        width={'size': 1,  "offset": 0}),
                    html.Div([
                        dbc.Checklist(id='white_background2',
                                    options = [{'value':'White Background','label':'White Background'}])
                    ], style={'padding':5}),
                    ],
                    style={'padding':5}),
                ],style={'padding':5}),
        ]),
        dbc.Row(),
        dbc.Row(),
        dbc.Row(),
        dbc.Row(),
        dbc.Row(
            dbc.Col(
                html.Div(id='container', children=[]),
                width={'size': 12,  "offset": 0}),style={"height": "300vh"}),
    ], style={'padding':20, 'text-align':'center'})

#---------------------------------------------------------------
app.layout = html.Div([
    dbc.Row(dbc.Col(html.H1("Analysis of indicators"),width={'size':12,"offset":3})),
    html.Div([
        dbc.Tabs(id='tabs-example', active_tab='About', className="nav nav-tabs", children=[
        dbc.Tab(label='About', tab_id='About', className='nav-link active'),
        dbc.Tab(label='Filter', tab_id='Filter', className='nav-link active'),
        dbc.Tab(label='Cluster', tab_id='Cluster', className='nav-link active'),
        dbc.Tab(label='Plot', tab_id='Plot', className='nav-link active'),
        ]),
        html.Div(id='tabs-example-content'),
    ])
])

#---------------------------------------------------------------
# Callback for the tab menu
@app.callback(Output('tabs-example-content', 'children'),
              Input('tabs-example', 'active_tab'))
def render_content(tab):
    if tab == 'About':
        return dbc.Row(dbc.Col(html.Div(children=[
                    html.P('This application allows you to compare composite indicators.'),
                    html.P('Steps to use the app:'),
                    html.P("1.    In the filter tab: Choose the countries you want to analize"),
                    html.P("2.    In the cluster tab: You can create clusters of countries according to indicators you select. You can visualize them graphically in some of the representations of the Plot Tab"),
                    html.P("3.    In the plot tab: Vizualize graphics of the parameters chosen before"),
                    ],
                style={'padding': 20}),
            width={'size': 12,  "offset":0})
        )
    elif tab == 'Filter':
        return filter_tab()
    elif tab== 'Cluster':
        return cluster_tab()
    elif tab== 'Plot':
        return plot_tab()

#Callback for the Buttons inside filter tab
@app.callback(
    Output(component_id='toast_saved', component_property='is_open'),
    [Input(component_id='save_considered', component_property='n_clicks'),
     State(component_id='ckl_considered', component_property='value')]
)

def update_graph(bt1,value):
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    counter=0
    if 'save_considered' in changed_id:
        a = (int(len(df_list_groups_of_countries.ISO_NUESTRO.values.tolist())))
        l_aux = [int(0)]*(int(len(df_list_groups_of_countries.ISO_NUESTRO.values.tolist())))
        if value: # Check if Checklist is empty
            Not_none_values = filter(None.__ne__, value)
            value= list(Not_none_values)
            value= sorted(value)
            for i in range(1,a):
                for counter in range(0,len(value)):
                    if(value[counter]==df_list_groups_of_countries.iat[i,2]):
                        l_aux[i]=1 # Assign one to selected countries
        df_aux=pd.DataFrame(l_aux)
        write_excel(df_aux,file2,'Hoja1',1,2)
        df_aux2 = pd.DataFrame(value)
        if not df_aux2.empty:
            df_aux2.columns = ["Last Saved Considered"]
        deleteColums("static/Grupos_paises.xlsx",'Hoja1',34,34)
        write_excel(pd.DataFrame(["Last Saved Considered"]),"static/Grupos_paises.xlsx",'Hoja1',0,34)
        write_excel(df_aux2,"static/Grupos_paises.xlsx","Hoja1",1,34)
        return True
    return False
@app.callback(
    Output(component_id='toast_desired', component_property='is_open'),
    [Input(component_id='save_desired', component_property='n_clicks'),
     Input(component_id='ckl_desired', component_property='value')]
)

def update_graph(bt1,value):
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    counter=0
    if 'save_desired' in changed_id:
        a = (int(len(df_list_groups_of_countries.ISO_NUESTRO.values.tolist())))
        l_aux = [int(0)]*(int(len(df_list_groups_of_countries.ISO_NUESTRO.values.tolist())))
        if value: # Check if Checklist is empty
            Not_none_values = filter(None.__ne__, value)
            value= list(Not_none_values)
            value= sorted(value)
            for i in range(1,a):
                for counter in range(0,len(value)):
                    if(value[counter]==df_list_groups_of_countries.iat[i,2]):
                        l_aux[i]=1 # Assign one to selected countries
        df_aux=pd.DataFrame(l_aux)
        write_excel(df_aux,file2,'Hoja1',1,3)
        df_aux2 = pd.DataFrame(value)
        if not df_aux2.empty:
            df_aux2.columns = ["Last Saved Desired"]
        deleteColums("static/Grupos_paises.xlsx",'Hoja1',35,35)
        write_excel(pd.DataFrame(["Last Saved Desired"]),"static/Grupos_paises.xlsx",'Hoja1',0,35)
        write_excel(df_aux2,"static/Grupos_paises.xlsx","Hoja1",1,35)
        return True
    return False




# Callbacks for the Dropdowns and buttons in the Filter tab

@app.callback(
    [Output(component_id='ckl_considered', component_property='value')],
    [Input(component_id='dropdown_groups_countries', component_property='value'),
    Input(component_id='include_all_considered', component_property='n_clicks'),
    Input(component_id='clear_all_considered', component_property='n_clicks')]
)
def update_dd1(dd,bt1,bt2):
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'include_all_considered' in changed_id:
        return [df_list_groups_of_countries['Include All'].values.tolist()]
    if 'clear_all_considered' in changed_id:
        return [df_list_groups_of_countries['Clear All'].values.tolist()]
    if 'dropdown_groups_countries' in changed_id:
        return [df_list_groups_of_countries[dd].values.tolist()]
    df_aux = pd.read_excel('static/Grupos_paises.xlsx',"Hoja1",engine='openpyxl')
    return [df_aux["Last Saved Considered"].values.tolist()]

@app.callback(
    [Output(component_id='number_countries_considered', component_property='children')],
    [Input(component_id='ckl_considered', component_property='value')]
)

def update_number_countries(ck1):
    Not_none_values = filter(None.__ne__, ck1)
    ck1= list(Not_none_values)
    return [str('Number of selected countries: ' + str(len(ck1)))]

    
@app.callback(
    [Output(component_id='ckl_desired', component_property='value')],
    [Input(component_id='dropdown_groups_countries_labels', component_property='value'),
    Input(component_id='include_all_desired', component_property='n_clicks'),
    Input(component_id='clear_all_desired', component_property='n_clicks')]
)
def update_dd1(dd,bt1,bt2):
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'include_all_desired' in changed_id:
        return [df_list_groups_of_countries['Include All'].values.tolist()]
    if 'clear_all_desired' in changed_id:
        return [df_list_groups_of_countries['Clear All'].values.tolist()]
    if 'dropdown_groups_countries_labels' in changed_id:
        return [df_list_groups_of_countries[dd].values.tolist()]
    df_aux = pd.read_excel('static/Grupos_paises.xlsx',"Hoja1",engine='openpyxl')
    return [df_aux["Last Saved Desired"].values.tolist()]


@app.callback(
[Output(component_id='number_countries_desired', component_property='children')],
[Input(component_id='ckl_desired', component_property='value')]
)

def update_number_countries(ck1):
    Not_none_values = filter(None.__ne__, ck1)
    ck1= list(Not_none_values)
    return [str('Number of selected countries: ' + str(len(ck1)))]

# Callbacks for the Dropdowns in the Cluster tab

# When you choose a different file or sheet you want to see the corresponding new sheets and indicators
@app.callback(
    [Output(component_id='dropdown_sheet_cluster', component_property='options'),
    Output(component_id='dropdown_sheet_cluster', component_property='value'), 
    Output(component_id='dropdown_indicator_cluster', component_property='options'),
    Output(component_id='dropdown_indicator_cluster', component_property='value')],
    [Input(component_id='dropdown_file_cluster', component_property='value'),
    Input(component_id='dropdown_sheet_cluster', component_property='value'),
    State(component_id='dropdown_sheet_cluster', component_property='options')]
)
def update_dd1(dd1,dd2,dd3):
    # Obtain dic with sheets from current selected file and a random value to
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'dropdown_file_cluster' in changed_id:
        wb=openpyxl.load_workbook(dd1)
        a=wb.sheetnames
        df_sheets_dict = pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_sheets = a[0]

        df_aux = pd.read_excel(dd1, sheet_name=value_sheets, engine='openpyxl')
        a=list(df_aux.columns)[2:]
        df_indicators_dict =pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_indicators = a[0]
        write_excel(pd.DataFrame([dd1,value_sheets,value_indicators]),file2,"ClusterDropdown",1,0)
        return df_sheets_dict, value_sheets, df_indicators_dict, value_indicators
    if 'dropdown_sheet_cluster' in changed_id:
        # Obtain dic with the indicators of the current selected sheet
        df_aux = pd.read_excel(dd1, sheet_name=dd2, engine='openpyxl')
        a=list(df_aux.columns)[2:]
        df_indicators_dict =pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_indicators = a[0]
        write_excel(pd.DataFrame([dd1,dd2,value_indicators]),file2,"ClusterDropdown",1,0)
        return dd3, dd2, df_indicators_dict, value_indicators
    return df_default_sheets_dict,"ODS",df_default_indicators_dict,'2020 Global Index Score (0-100)'


# Callbacks for the Dropdowns in the Plot tab

# Plot 2D: x axis
# When you choose a different file or sheet you want to see the corresponding new sheets and indicators
@app.callback(
    [Output(component_id='dropdown_sheet_x', component_property='options'),
    Output(component_id='dropdown_sheet_x', component_property='value'), 
    Output(component_id='dropdown_indicator_x', component_property='options'),
    Output(component_id='dropdown_indicator_x', component_property='value')],
    [Input(component_id='dropdown_file_x', component_property='value'),
    Input(component_id='dropdown_sheet_x', component_property='value'),
    State(component_id='dropdown_sheet_x', component_property='options')]
)
def update_dd1(dd1,dd2,dd3):
    # Obtain dic with sheets from current selected file and a random value to
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'dropdown_file_x' in changed_id:
        wb=openpyxl.load_workbook(dd1)
        a=wb.sheetnames
        df_sheets_dict = pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_sheets = a[0]

        df_aux = pd.read_excel(dd1, sheet_name=value_sheets, engine='openpyxl')
        a=list(df_aux.columns)[2:]
        df_indicators_dict =pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_indicators = a[0]
        return df_sheets_dict, value_sheets, df_indicators_dict, value_indicators
    if 'dropdown_sheet_x' in changed_id:
        # Obtain dic with the indicators of the current selected sheet
        df_aux = pd.read_excel(dd1, sheet_name=dd2, engine='openpyxl')
        a=list(df_aux.columns)[2:]
        df_indicators_dict =pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_indicators = a[0]
        return dd3, dd2, df_indicators_dict, value_indicators
    
    return df_default_sheets_dict,"ODS",df_default_indicators_dict,'2020 Global Index Score (0-100)'

# Plot 2D: y axis
# When you choose a different file or sheet you want to see the corresponding new sheets and indicators
@app.callback(
    [Output(component_id='dropdown_sheet_y', component_property='options'),
    Output(component_id='dropdown_sheet_y', component_property='value'), 
    Output(component_id='dropdown_indicator_y', component_property='options'),
    Output(component_id='dropdown_indicator_y', component_property='value')],
    [Input(component_id='dropdown_file_y', component_property='value'),
    Input(component_id='dropdown_sheet_y', component_property='value'),
    State(component_id='dropdown_sheet_y', component_property='options')]
)
def update_dd1(dd1,dd2,dd3):
    # Obtain dic with sheets from current selected file and a random value to
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'dropdown_file_y' in changed_id:
        wb=openpyxl.load_workbook(dd1)
        a=wb.sheetnames
        df_sheets_dict = pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_sheets = a[0]

        df_aux = pd.read_excel(dd1, sheet_name=value_sheets, engine='openpyxl')
        a=list(df_aux.columns)[2:]
        df_indicators_dict =pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_indicators = a[0]
        return df_sheets_dict, value_sheets, df_indicators_dict, value_indicators
    if 'dropdown_sheet_y' in changed_id:
        # Obtain dic with the indicators of the current selected sheet
        df_aux = pd.read_excel(dd1, sheet_name=dd2, engine='openpyxl')
        a=list(df_aux.columns)[2:]
        df_indicators_dict =pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_indicators = a[0]
        return dd3, dd2, df_indicators_dict, value_indicators
    
    return df_default_sheets_dict,"ODS",df_default_indicators_dict,'2020 Global Index Score (0-100)'

# Sunburst
# When you choose a different file or sheet you want to see the corresponding new sheets and indicators
@app.callback(
    [Output(component_id='dropdown_sheet_sunburst', component_property='options'),
    Output(component_id='dropdown_sheet_sunburst', component_property='value'), 
    Output(component_id='dropdown_indicator_sunburst', component_property='options'),
    Output(component_id='dropdown_indicator_sunburst', component_property='value')],
    [Input(component_id='dropdown_file_sunburst', component_property='value'),
    Input(component_id='dropdown_sheet_sunburst', component_property='value'),
    State(component_id='dropdown_sheet_sunburst', component_property='options')]
)
def update_dd1(dd1,dd2,dd3):
    # Obtain dic with sheets from current selected file and a random value to
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'dropdown_file_sunburst' in changed_id:
        wb=openpyxl.load_workbook(dd1)
        a=wb.sheetnames
        df_sheets_dict = pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_sheets = a[0]

        df_aux = pd.read_excel(dd1, sheet_name=value_sheets, engine='openpyxl')
        a=list(df_aux.columns)[2:]
        df_indicators_dict =pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_indicators = a[0]
        return df_sheets_dict, value_sheets, df_indicators_dict, value_indicators
    if 'dropdown_sheet_sunburst' in changed_id:
        # Obtain dic with the indicators of the current selected sheet
        df_aux = pd.read_excel(dd1, sheet_name=dd2, engine='openpyxl')
        a=list(df_aux.columns)[2:]
        df_indicators_dict =pd.DataFrame({'label':a,'value':a}).to_dict('records')
        value_indicators = a[0]
        return dd3, dd2, df_indicators_dict, value_indicators
    
    return df_default_sheets_dict,"ODS",df_default_indicators_dict,'2020 Global Index Score (0-100)'

# # Add a variable to the cluster list

@app.callback(
    [Output(component_id='list_variables_cluster', component_property='children')],
    [Input(component_id='confirm_variable_cluster', component_property='n_clicks'),
    Input(component_id='clear_variable_list', component_property='n_clicks'),
    State(component_id='dropdown_file_cluster', component_property='value'),
    State(component_id='dropdown_sheet_cluster', component_property='value'),
    State(component_id='dropdown_indicator_cluster', component_property='value'),
    State('list_variables_cluster','children')]
)

def update_list_cluster(bt1,bt2,file_chosen,sheet_chosen,indicator_chosen, listChildren):
    
    ctx = dash.callback_context
    if not ctx.triggered:
        raise dash.exceptions.PreventUpdate
    else:
        button_id = ctx.triggered[0]['prop_id'].split('.')[0]

    if button_id not in ['confirm_variable_cluster', 'clear_variable_list']:
        raise dash.exceptions.PreventUpdate
    else:
        if not listChildren:
            #df_for_list_cluster_variable = pd.DataFrame()
            listChildren = []
            #df_aux = pd.read_excel(file2, sheet_name="Cluster_var", engine='openpyxl')
            #counter = 0
            #if df_aux.size==0:
            #    while(len(df_aux.iat[2,counter])>1):
            #        listChildren.append(dbc.ListGroupItem(df_aux.iat[2,counter]))
            #        counter = counter + 1
        if button_id == 'clear_variable_list':
            n=0
            deleteSheet(file2,'Cluster_var')
            count_number_of_list_cluster.iat[0,0]=0
            clear_dataframe = pd.DataFrame(columns=['var1','var2','var3','var4','var5','var6','var7','var8','var9','var10','var11','var12'])
            write_excel2(clear_dataframe,file2,"Cluster_var",0,0)
            return [[]]
        listChildren.append(dbc.ListGroupItem(indicator_chosen))
        # cluster_var = count_number_of_list_clusterpd.read_excel('static/filtros.xlsx', sheet_name='Cluster_var', engine='openpyxl')
        n= count_number_of_list_cluster.iat[0,0]
        count_number_of_list_cluster.iat[0,0]=(int(n)+1)
        # df5 = pd.DataFrame([int(n+1)], columns=list('n'))
        # write_excel(df5,'static/filtros.xlsx','Cluster_var',0,0)
        write_excel(pd.DataFrame([file_chosen,sheet_chosen,indicator_chosen]),file2,"Cluster_var",1,int(n))
        #df_for_list_cluster_variable = df_for_list_cluster_variable.append([file_chosen,sheet_chosen,indicator_chosen])

    
    return [listChildren]

    # Show the list of countries that belong to a cluster
    
@app.callback(
    [Output(component_id='future_table_of_clusters', component_property='children')],
    [Input(component_id='make_cluster_button', component_property='n_clicks'),
    State(component_id='list_variables_cluster', component_property='value'),
    State(component_id='dropdown_number_clusters', component_property='value')]
)

def show_list_of_clusters(cluster_button,list_variables_cluster,number_clusters):
    # DataFrame with file, sheet and indicator of cluster variables
    # if df_cluster_vars:
    #     return True
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]

    if 'make_cluster_button' in changed_id:
        # return dbc.Table(
        # [html.Thead(html.Tr([html.Th("Cluster 1"), html.Th("Cluster 2"), html.Th("Cluster 3"), html.Th("Cluster 4"),  html.Th("Cluster 5")]))],
        # id ='table_of_clusters'
        df_cluster_vars_info = pd.read_excel(file2, sheet_name='Cluster_var', engine='openpyxl')
        df_cluster_vars = read_indicator(df_cluster_vars_info.iat[0,0],df_cluster_vars_info.iat[1,0],df_cluster_vars_info.iat[2,0])
        counter_aux = 1
        while(len(str(df_cluster_vars_info.iat[0,counter_aux]))>3):
            next_cluster_var = read_indicator(df_cluster_vars_info.iat[0,counter_aux],df_cluster_vars_info.iat[1,counter_aux],df_cluster_vars_info.iat[2,counter_aux])
            df_cluster_vars = pd.concat([df_cluster_vars, next_cluster_var], axis=1)
            counter_aux = counter_aux + 1
        #df_cluster_vars = df_cluster_vars.fillna(0)
        df1 = pd.read_excel(file2, sheet_name='Hoja1', engine='openpyxl')

        deleteColums(file2,'Cluster',4,20)
        write_excel2(df_cluster_vars,file2,'Cluster',0,4)

        df_cluster_vars_for_parallel = pd.DataFrame()
        for i in range(int(df_cluster_vars.shape[1])):
            #a= df_cluster_vars[df_cluster_vars.columns[i]].astype(str)
            df_cluster_vars_for_parallel[df_cluster_vars_info.iat[2,i]] = np.round(df_cluster_vars[df_cluster_vars.columns[i]],2).astype(str) + ' ' + df1["country"]
        
        for i in range(int(df_cluster_vars.shape[0])):
            if (df_cluster_vars.iloc[i].isnull().values.any()) | (df1.Considerar_pais.iloc[i]==0):
                df_cluster_vars_for_parallel.iloc[i]=np.nan

        # for i in range(int(df_cluster_vars.shape[0])):    
        #     if df_cluster_vars.iloc[i].isnull().values.any():
        #         df_cluster_vars_for_parallel.loc[i]=np.nan
        #     else:
        
        deleteSheet(file2,'Parallel')
        write_excel2(df_cluster_vars_for_parallel,file2,'Parallel',0,4)


        df_cluster_vars_considered = df_cluster_vars[df1.Considerar_pais != 0]

        df_cluster_vars_without_nan = df_cluster_vars_considered.dropna()
        kmeans = KMeans(n_clusters=int(number_clusters), random_state=0).fit(df_cluster_vars_without_nan)
        index_cluster = kmeans.labels_

        index_cluster_with_nan_only_considered = pd.DataFrame(columns=['Cluster'])
        counter2 = 0
        for i in range(int(df_cluster_vars_considered.shape[0])):
            if df_cluster_vars_considered.iloc[i].isnull().values.any():
                index_cluster_with_nan_only_considered.loc[i]=np.nan
            else:
                index_cluster_with_nan_only_considered.loc[i]=index_cluster.item(counter2)
                counter2 = counter2 + 1

        index_cluster_with_nan = pd.DataFrame(columns=['Cluster'])
        counter2 = 0
        for i in range(int(df_cluster_vars.shape[0])):
            if df1.Considerar_pais.iloc[i]==0:
                index_cluster_with_nan.loc[i]=np.nan
            else:
                index_cluster_with_nan.loc[i]=index_cluster_with_nan_only_considered.loc[counter2]
                counter2 = counter2 + 1






        # I am here: the indexes of the clusters are obtained by there are two things that I do not do:
        # considerar country is not taken into account here (it takes all the countries it can)
        # since I have removed the rows with nan, now I do not know which row of index_cluster corresponds to
        # each country, and therefore, I do not know how to write the matrix in the excel










        write_excel2(index_cluster_with_nan,file2,'Cluster',0,2)
        write_excel2(index_cluster_with_nan,file2,'Parallel',0,2)
        # Represent the table once the clusters are written in the excel
        dfClustersSheet = pd.read_excel('static/filtros.xlsx', sheet_name='Cluster', engine='openpyxl')
        dfClustersSheet2 = dfClustersSheet[["country","Cluster"]]
        n=1
        new_column = dfClustersSheet2["country"].loc[dfClustersSheet2['Cluster']==0]
        new_column.reset_index(drop=True, inplace=True)
        df_total = new_column
        while(dfClustersSheet2["country"].loc[dfClustersSheet2['Cluster']==n].empty==False):
            new_column = dfClustersSheet2["country"].loc[dfClustersSheet2['Cluster']==n]
            new_column.reset_index(drop=True, inplace=True)
            new_column.columns='Cluster'+str(n)
            #new_column.header(0)= 'Cluster '+str(n)
            df_total = pd.concat([df_total, new_column], ignore_index=True, axis=1)
            #df_total.loc[:,'Cluster '+str(n)]=new_column
            #df_total['Cluster '+str(n)]=new_column
            #df_total = df_total.rename(index={0: 'Cluster'+str(n)})
            n=n+1
        deleteColums(file2,"Cluster_table",0,10)
        write_excel2(df_total,file2,'Cluster_table',0,0)
        return [dbc.Table.from_dataframe(df_total, striped=True,bordered = True, hover=True)]
    df_aux = pd.read_excel(file2, sheet_name='Cluster_table', engine='openpyxl')
    return [dbc.Table.from_dataframe(df_aux, striped=True,bordered = True, hover=True)]

# Callback for the Buttons inside Plot tab
@app.callback(
    [Output(component_id='container', component_property='children')],
    [Input(component_id='submit_parallel', component_property='n_clicks'),
    Input(component_id='submit_radar', component_property='n_clicks'),
    Input(component_id='submit_map', component_property='n_clicks'),
    Input(component_id='submit_sunburst', component_property='n_clicks'),
    Input(component_id='submit_vertex', component_property='n_clicks'),
    Input(component_id='submit_vertex3d', component_property='n_clicks'),
    Input(component_id='submit_plot2D', component_property='n_clicks'),
    State(component_id='dropdown_file_sunburst', component_property='value'),
    State(component_id='dropdown_sheet_sunburst', component_property='value'),
    State(component_id='dropdown_indicator_sunburst', component_property='value'),
    State(component_id='dropdown_file_x', component_property='value'),
    State(component_id='dropdown_sheet_x', component_property='value'),
    State(component_id='dropdown_indicator_x', component_property='value'),
    State(component_id='dropdown_file_y', component_property='value'),
    State(component_id='dropdown_sheet_y', component_property='value'),
    State(component_id='dropdown_indicator_y', component_property='value'),
    State(component_id='method_vertex', component_property='value'),
    State(component_id='dropdown_neighbors_vertex', component_property='value'),
    State(component_id='ckl_types_correlations', component_property='value'),
    State(component_id='white_background', component_property='value'),
    State(component_id='white_background2', component_property='value')
    ]
    )

def update_graph(bt1,bt2,bt3,bt4,bt5,bt6,bt7,dd1_sun,dd2_sun,dd3_sun,dd1_2d,dd2_2d,dd3_2d,dd4_2d,dd5_2d,dd6_2d,method_vertex,neigh_vertex,types_correlations,white_back,white_back2):
    changed_id = [p['prop_id'] for p in dash.callback_context.triggered][0]
    if 'submit_parallel' in changed_id:
        df1 = pd.read_excel(file2, sheet_name='Hoja1', engine='openpyxl')
        df2 = pd.read_excel(file2, sheet_name='Cluster', engine='openpyxl')
        df3 = pd.read_excel(file2, sheet_name='Parallel', engine='openpyxl')
        new_child = [html.Div(
            style={'padding': 20, 'text-align':'center', 'display': 'inline-block'},
            children=[
                dcc.Graph(
                    id= 'radar',
                    figure=draw_parallel(df1,df2,df3)
                    )
                ]
                )]
        return new_child
    elif 'submit_radar' in changed_id:
        df1 = pd.read_excel(file2, sheet_name='Hoja1', engine='openpyxl')
        df2 = pd.read_excel(file2, sheet_name='Cluster', engine='openpyxl')
        new_child = [html.Div(
            style={'padding': 20, 'text-align':'center', 'display': 'inline-block'},
            children=[
                dcc.Graph(
                    id= 'radar',
                    figure=draw_radar(df1,df2, white_back2)
                    )
                ]
                )]
        return new_child

    elif 'submit_sunburst' in changed_id:
        df1 = pd.read_excel(file2, sheet_name='Hoja1', engine='openpyxl')
        df2 = pd.read_excel(file2, sheet_name='Cluster', engine='openpyxl')
        df4 = read_indicator_sunburst(dd1_sun,dd2_sun,dd3_sun)
        new_child = [html.Div(
            style={'padding': 20, 'text-align':'center', 'display': 'inline-block'},
            children=[
                dcc.Graph(
                    id= 'radar',
                    figure=draw_sunburst(df1,df2,df4)
                    )
                ]
                )]
        return new_child
        #return [html.Div(id='sth')]

    elif 'submit_map' in changed_id:
        df1 = pd.read_excel(file2, sheet_name='Hoja1', engine='openpyxl')
        df2 = pd.read_excel(file2, sheet_name='Cluster', engine='openpyxl')
        new_child = [html.Div(
            style={'padding': 100, 'text-align':'center', 'display': 'inline-block'},
            children=[
                dcc.Graph(
                    id= 'radar',
                    figure=draw_map(df1,df2)
                    )
                ])]
        return new_child
        
    elif ('submit_vertex' in changed_id) &  (('submit_vertex3d' in changed_id)==False):
        #df5 = pd.read_excel(file2, sheet_name='Vertex', engine='openpyxl')
        #write_excel([df_aux],file2,'Vertex',2,1)
        df1 = pd.read_excel(file2, sheet_name='Hoja1', engine='openpyxl')
        df2 = pd.read_excel(file2, sheet_name='Cluster', engine='openpyxl')
        df5 = pd.DataFrame([[method_vertex,neigh_vertex]], columns=list('AB'))
        new_child = [html.Div(
            style={'padding': 100, 'text-align':'center', 'display': 'inline-block'},
            children=[
                dcc.Graph(
                    id= 'vertex',
                    figure=draw_vertex(df1,df2,df5)
                    )
                ])]
        return new_child

    elif ('submit_vertex3d' in changed_id):
        df1 = pd.read_excel(file2, sheet_name='Hoja1', engine='openpyxl')
        df2 = pd.read_excel(file2, sheet_name='Cluster', engine='openpyxl')
        df5 = pd.DataFrame([[method_vertex,neigh_vertex]], columns=list('AB'))
        new_child = [html.Div(
            style={'padding': 100, 'text-align':'center', 'display': 'inline-block'},
            children=[
                dcc.Graph(
                    id= 'vertex3d',
                    figure=draw_vertex3d(df1,df2,df5)
                    )
                ])]
        return new_child
    elif 'submit_plot2D' in changed_id:
        new_child = [html.Div(
            style={'padding': 100, 'text-align':'center', 'display': 'inline-block'},
            children=[
                dcc.Graph(
                    id= 'radar',
                    figure=plot_2d(dd1_2d,dd2_2d,dd3_2d,dd4_2d,dd5_2d,dd6_2d,types_correlations, white_back)
                    )
                ])]
        return new_child

    return [html.Div(id='sth')]
    
if __name__ == '__main__':
    app.run_server(debug=True,dev_tools_ui=True, dev_tools_props_check=True)